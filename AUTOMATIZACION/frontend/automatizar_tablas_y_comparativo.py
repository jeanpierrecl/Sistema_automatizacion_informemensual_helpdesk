from __future__ import annotations

import argparse
import calendar
import copy
import csv
from datetime import date
import re
import subprocess
import zipfile
import xml.etree.ElementTree as ET
from collections import defaultdict
from pathlib import Path
from typing import Any
import uuid

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
NS = "{%s}" % NS_MAIN
NS_R = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
NS_MC = "http://schemas.openxmlformats.org/markup-compatibility/2006"
KNOWN_PREFIX_NAMESPACES = {
    "x14ac": "http://schemas.microsoft.com/office/spreadsheetml/2009/9/ac",
    "xr": "http://schemas.microsoft.com/office/spreadsheetml/2014/revision",
    "xr2": "http://schemas.microsoft.com/office/spreadsheetml/2015/revision2",
    "xr3": "http://schemas.microsoft.com/office/spreadsheetml/2016/revision3",
    "x15": "http://schemas.microsoft.com/office/spreadsheetml/2010/11/main",
    "xr6": "http://schemas.microsoft.com/office/spreadsheetml/2016/revision6",
    "xr10": "http://schemas.microsoft.com/office/spreadsheetml/2016/revision10",
}
ET.register_namespace("", NS_MAIN)
ET.register_namespace("r", NS_R)
ET.register_namespace("mc", NS_MC)
for _prefix, _uri in KNOWN_PREFIX_NAMESPACES.items():
    ET.register_namespace(_prefix, _uri)
SHARED_STRINGS_CONTEXT = None

MESES = {
    "ene": 1,
    "feb": 2,
    "mar": 3,
    "abr": 4,
    "may": 5,
    "jun": 6,
    "jul": 7,
    "ago": 8,
    "sep": 9,
    "oct": 10,
    "nov": 11,
    "dic": 12,
}

MESES_ABREV = {
    1: "Ene",
    2: "Feb",
    3: "Mar",
    4: "Abr",
    5: "May",
    6: "Jun",
    7: "Jul",
    8: "Ago",
    9: "Sep",
    10: "Oct",
    11: "Nov",
    12: "Dic",
}


EXCEL_COLOR_D9E1F2 = 15917529
EXCEL_COLOR_D9D9D9 = 14277081
CURRENCY_FORMAT = '[$$-540A]#,##0.00'
ACCOUNTING_FORMAT = CURRENCY_FORMAT


def ps_string(valor: str) -> str:
    return str(valor).replace("'", "''")


def es_relleno_amarillo(cell) -> bool:
    fill = getattr(cell, "fill", None)
    if not fill or fill.fill_type is None:
        return False
    color = fill.fgColor
    if color is None:
        return False
    if color.type == "rgb" and color.rgb:
        return color.rgb.upper().endswith("FFFF00")
    if color.type == "indexed" and color.indexed == 6:
        return True
    return False


def celdas_amarillas_por_header(ruta_instances: str | Path | None) -> list[tuple[int, str]]:
    if not ruta_instances:
        return []
    ruta = Path(ruta_instances)
    if not ruta.exists():
        return []
    wb = load_workbook(ruta, data_only=False, read_only=False)
    try:
        ws = wb[wb.sheetnames[0]]
        headers = {cell.column: str(cell.value).strip() for cell in ws[1] if cell.value not in (None, "")}
        marcas: list[tuple[int, str]] = []
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                header = headers.get(cell.column)
                if header and es_relleno_amarillo(cell):
                    marcas.append((cell.row, header))
        return marcas
    finally:
        wb.close()


def reparar_con_excel_si_disponible(
    ruta_excel: str | Path,
    ruta_instances: str | Path | None = None,
    etiqueta_mes: str | None = None,
    accion_mes: str = "auto",
) -> bool:
    ruta = Path(ruta_excel).resolve()
    temporal = ruta.with_name(f"{ruta.stem}_excel_reparado_{uuid.uuid4().hex[:8]}{ruta.suffix}")
    instances_path = str(Path(ruta_instances).resolve()).replace("'", "''") if ruta_instances else ""
    etiqueta_mes_ps = (etiqueta_mes or "").replace("'", "''")
    accion_mes_ps = (accion_mes or "auto").replace("'", "''")
    marcas_amarillas = celdas_amarillas_por_header(ruta_instances)
    marcas_amarillas_path = ruta.with_name(f"{ruta.stem}_amarillos_{uuid.uuid4().hex[:8]}.csv")
    if marcas_amarillas:
        with marcas_amarillas_path.open("w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow(["Row", "Header"])
            writer.writerows(marcas_amarillas)
    marcas_amarillas_csv_ps = str(marcas_amarillas_path).replace("'", "''") if marcas_amarillas else ""
    ps_script = f"""
$ErrorActionPreference = 'Stop'
$excel = New-Object -ComObject Excel.Application
$excel.Visible = $false
$excel.DisplayAlerts = $false
try {{
  $wb = $excel.Workbooks.Open('{str(ruta).replace("'", "''")}', 0, $false, 5, '', '', $true, 1, '', $false, $false, 0, $false, $true, 1)
  foreach ($ws in $wb.Worksheets) {{
    try {{ $ws.Cells.Replace('(en blanco)', '', 1, 1, $false, $false, $false, $false) | Out-Null }} catch {{}}
  }}
  $instancesPath = '{instances_path}'
  $monthLabelParam = '{etiqueta_mes_ps}'
  $monthAction = '{accion_mes_ps}'
  if ($instancesPath -ne '' -and $monthLabelParam -ne '') {{
    $monthParts = $monthLabelParam.Split('-')
    $monthMap = @{{'Ene'=1;'Feb'=2;'Mar'=3;'Abr'=4;'May'=5;'Jun'=6;'Jul'=7;'Ago'=8;'Sep'=9;'Oct'=10;'Nov'=11;'Dic'=12}}
    if ($monthParts.Count -eq 2 -and $monthMap.ContainsKey($monthParts[0])) {{
      $monthDateForDetail = [datetime]::new((2000 + [int]$monthParts[1]), $monthMap[$monthParts[0]], 1)
      $detail = $wb.Worksheets.Item('Detalle Ene-Mar')
      $srcWb = $excel.Workbooks.Open($instancesPath, 0, $true)
      try {{
        $src = $srcWb.Worksheets.Item(1)
        $srcUsed = $src.UsedRange
        $srcLastRow = $srcUsed.Row + $srcUsed.Rows.Count - 1
        $srcLastCol = $srcUsed.Column + $srcUsed.Columns.Count - 1
        $detailUsed = $detail.UsedRange
        $detailLastCol = $detailUsed.Column + $detailUsed.Columns.Count - 1
        $srcHeaders = @{{}}
        for ($c = 1; $c -le $srcLastCol; $c++) {{
          $h = [string]$src.Cells.Item(1, $c).Text
          if ($h -ne '') {{ $srcHeaders[$h] = $c }}
        }}
        if ($monthAction -eq 'actualizar') {{
          $detailLastRowForDelete = $detail.UsedRange.Row + $detail.UsedRange.Rows.Count - 1
          for ($r = $detailLastRowForDelete; $r -ge 2; $r--) {{
            $v = $detail.Cells.Item($r, 1).Value2
            if ($v -is [double] -or $v -is [int]) {{
              $d = [DateTime]::FromOADate([double]$v)
              if ($d.Year -eq $monthDateForDetail.Year -and $d.Month -eq $monthDateForDetail.Month) {{
                $detail.Rows.Item($r).Delete() | Out-Null
              }}
            }}
          }}
        }}
        $destRow = $detail.UsedRange.Row + $detail.UsedRange.Rows.Count
        $rowCount = $srcLastRow - 1
        if ($rowCount -gt 0) {{
          $srcValues = $src.Range($src.Cells.Item(1, 1), $src.Cells.Item($srcLastRow, $srcLastCol)).Value2
          $outputValues = [System.Array]::CreateInstance([object], @($rowCount, $detailLastCol), @(1, 1))
          $monthSerial = $monthDateForDetail.ToOADate()
          $detailHeaders = @{{}}
          $detailHeaderCols = @{{}}
          for ($c = 1; $c -le $detailLastCol; $c++) {{
            $headerText = [string]$detail.Cells.Item(1, $c).Text
            $detailHeaders[$c] = $headerText
            if ($headerText -ne '' -and -not $detailHeaderCols.ContainsKey($headerText)) {{
              $detailHeaderCols[$headerText] = $c
            }}
          }}
          for ($r = 1; $r -le $rowCount; $r++) {{
            $srcRow = $r + 1
            $outputValues.SetValue($monthSerial, $r, 1)
            for ($c = 2; $c -le $detailLastCol; $c++) {{
              $header = $detailHeaders[$c]
              if ($srcHeaders.ContainsKey($header)) {{
                $srcValue = $srcValues[$srcRow, $srcHeaders[$header]]
                if ($srcValue -is [bool]) {{
                  $srcValue = [int]$srcValue
                }}
                if ($header -eq 'Instance ID' -and $null -ne $srcValue) {{
                  $srcValue = [string]$srcValue
                }}
                $outputValues.SetValue($srcValue, $r, $c)
              }}
            }}
          }}
          $target = $detail.Range($detail.Cells.Item($destRow, 1), $detail.Cells.Item($destRow + $rowCount - 1, $detailLastCol))
          for ($c = 1; $c -le $detailLastCol; $c++) {{
            if ([string]$detail.Cells.Item(1, $c).Text -eq 'Instance ID') {{
              $detail.Range($detail.Cells.Item($destRow, $c), $detail.Cells.Item($destRow + $rowCount - 1, $c)).NumberFormat = '@'
            }}
          }}
          $target.Value2 = $outputValues
          $detail.Range($detail.Cells.Item($destRow, 1), $detail.Cells.Item($destRow + $rowCount - 1, 1)).NumberFormat = 'mmm-yy'
          $highlightedPath = '{marcas_amarillas_csv_ps}'
          $highlightedCells = @()
          if ($highlightedPath -ne '' -and (Test-Path $highlightedPath)) {{
            $highlightedCells = Import-Csv -Path $highlightedPath
          }}
          $highlightedApplied = 0
          foreach ($mark in $highlightedCells) {{
            $header = [string]$mark.Header
            if ($detailHeaderCols.ContainsKey($header)) {{
              $destExcelRow = $destRow + [int]$mark.Row - 2
              $destExcelCol = [int]$detailHeaderCols[$header]
              $detail.Cells.Item($destExcelRow, $destExcelCol).Interior.Color = 65535
              $highlightedApplied++
            }}
          }}
          if ($highlightedApplied -gt 0) {{
            Write-Host "[Excel] Celdas amarillas copiadas al detalle: $highlightedApplied"
          }}
        }}
        $detailLastRow = $detail.UsedRange.Row + $detail.UsedRange.Rows.Count - 1
        $sourceData = $detail.Range($detail.Cells.Item(1, 1), $detail.Cells.Item($detailLastRow, $detailLastCol))
        foreach ($pivotSheetName in @('Variacion Barel Metal', 'VS&AS')) {{
          $pivotSheet = $wb.Worksheets.Item($pivotSheetName)
          foreach ($pt in $pivotSheet.PivotTables()) {{
            $cache = $wb.PivotCaches().Create(1, $sourceData)
            $pt.ChangePivotCache($cache)
            try {{ $pt.PivotFields('Mes').ClearAllFilters() | Out-Null }} catch {{}}
            $pt.RefreshTable() | Out-Null
          }}
        }}
      }} finally {{
        $srcWb.Close($false)
      }}
    }}
  }}
  $resumen = $wb.Worksheets.Item('Resumen')
  $used = $resumen.UsedRange
  $lastRow = $used.Row + $used.Rows.Count - 1
  $lastCol = $used.Column + $used.Columns.Count - 1
  $totalVarCol = 0
  $totalGeneralCol = 0
  $totalGeneralRow = 0
  $monthTotalCols = @()
  for ($c = 1; $c -le $lastCol; $c++) {{
    $header = [string]$resumen.Cells.Item(4, $c).Text
    if ($header -eq 'Total Var.') {{ $totalVarCol = $c }}
    if ($header -eq 'Total general') {{ $totalGeneralCol = $c }}
    if ($header.StartsWith('Total ') -and $header -ne 'Total general' -and $header -ne 'Total Var.') {{ $monthTotalCols += $c }}
  }}
  for ($r = 1; $r -le $lastRow; $r++) {{
    if ([string]$resumen.Cells.Item($r, 1).Text -eq 'Total general') {{ $totalGeneralRow = $r }}
  }}
  if ($totalVarCol -gt 0 -and $totalGeneralRow -gt 0 -and $monthTotalCols.Count -ge 2) {{
    $monthTotalCols = $monthTotalCols | Sort-Object
    $currentTotalCol = $monthTotalCols[$monthTotalCols.Count - 1]
    $previousTotalCol = $monthTotalCols[$monthTotalCols.Count - 2]
    $currentStartCol = $currentTotalCol - 6
    $previousStartCol = $previousTotalCol - 6
    $variationStartCol = $totalVarCol - 6
    $variationDataRange = $resumen.Range($resumen.Cells.Item(6, $variationStartCol), $resumen.Cells.Item($totalGeneralRow + 1, $totalVarCol))
    $variationDataRange.Interior.ColorIndex = -4142
    $variationDataRange.ClearContents() | Out-Null
    try {{ $variationDataRange.FormatConditions.Delete() | Out-Null }} catch {{}}
    for ($r = 6; $r -le $totalGeneralRow; $r++) {{
      for ($c = $variationStartCol; $c -le $totalVarCol; $c++) {{
        $offset = $c - $variationStartCol
        $currentRaw = $resumen.Cells.Item($r, $currentStartCol + $offset).Value2
        $previousRaw = $resumen.Cells.Item($r, $previousStartCol + $offset).Value2
        $currentValue = 0.0
        $previousValue = 0.0
        if ($currentRaw -is [double] -or $currentRaw -is [int]) {{ $currentValue = [double]$currentRaw }}
        if ($previousRaw -is [double] -or $previousRaw -is [int]) {{ $previousValue = [double]$previousRaw }}
        $value = $currentValue - $previousValue
        if ([math]::Abs([double]$value) -lt 0.005) {{
          $resumen.Cells.Item($r, $c).Value2 = '-'
          $resumen.Cells.Item($r, $c).Interior.ColorIndex = -4142
        }} else {{
          $resumen.Cells.Item($r, $c).Value2 = $value
          if ($value -gt 0) {{
            $resumen.Cells.Item($r, $c).Interior.Color = 255
          }} elseif ($value -lt 0) {{
            $resumen.Cells.Item($r, $c).Interior.Color = 5296274
          }}
        }}
      }}
    }}
    $currentTotalValue = 0.0
    $previousTotalValue = 0.0
    $currentTotalRaw = $resumen.Cells.Item($totalGeneralRow, $currentTotalCol).Value2
    $previousTotalRaw = $resumen.Cells.Item($totalGeneralRow, $previousTotalCol).Value2
    if ($currentTotalRaw -is [double] -or $currentTotalRaw -is [int]) {{ $currentTotalValue = [double]$currentTotalRaw }}
    if ($previousTotalRaw -is [double] -or $previousTotalRaw -is [int]) {{ $previousTotalValue = [double]$previousTotalRaw }}
    $monthlyDiff = $currentTotalValue - $previousTotalValue
    if ([math]::Abs($monthlyDiff) -lt 0.005 -or [math]::Abs($previousTotalValue) -lt 0.005) {{
      $resumen.Cells.Item($totalGeneralRow + 1, $totalVarCol).Value2 = '-'
    }} else {{
      $resumen.Cells.Item($totalGeneralRow + 1, $totalVarCol).Value2 = [math]::Abs($monthlyDiff / $previousTotalValue * 100)
    }}
    $cuadro = $wb.Worksheets.Item('Cuadro')
    $monthLabel = ([string]$resumen.Cells.Item(4, $currentTotalCol).Text).Replace('Total ', '')
    $cuadroLastMonthCol = 2
    $cuadroMonthCol = 0
    for ($c = 2; $c -le 60; $c++) {{
      $text = [string]$cuadro.Cells.Item(4, $c).Text
      if ($text -match '^[A-Z][a-z]{{2}}-\\d{{2}}$') {{
        $cuadroLastMonthCol = $c
        if ($text -eq $monthLabel) {{ $cuadroMonthCol = $c }}
      }}
    }}
    if ($cuadroMonthCol -eq 0) {{ $cuadroMonthCol = $cuadroLastMonthCol + 1 }}
    $cuadroDailyCol = $cuadroMonthCol - 1
    $currencyFormat = '{CURRENCY_FORMAT}'
    $accountingFormat = '{ACCOUNTING_FORMAT}'
    $headerColor = {EXCEL_COLOR_D9E1F2}
    $totalColor = {EXCEL_COLOR_D9D9D9}
    $cuadro.Cells.Item(4, $cuadroMonthCol).Value = $monthLabel
    $devValue = [double]$resumen.Cells.Item($totalGeneralRow, $currentStartCol).Value2
    $devqaValue = [double]$resumen.Cells.Item($totalGeneralRow, $currentStartCol + 1).Value2
    $qaValue = [double]$resumen.Cells.Item($totalGeneralRow, $currentStartCol + 2).Value2
    $prodValue = [double]$resumen.Cells.Item($totalGeneralRow, $currentStartCol + 4).Value2
    $allValue = [double]$resumen.Cells.Item($totalGeneralRow, $currentStartCol + 5).Value2
    $cuadro.Cells.Item(5, $cuadroMonthCol).Value2 = $devValue + ($devqaValue / 2) + ($allValue / 3)
    $cuadro.Cells.Item(6, $cuadroMonthCol).Value2 = $qaValue + ($devqaValue / 2) + ($allValue / 3)
    $cuadro.Cells.Item(7, $cuadroMonthCol).Value2 = $prodValue + ($allValue / 3)
    $cuadro.Cells.Item(8, $cuadroMonthCol).Formula = '=SUM(' + $cuadro.Cells.Item(5, $cuadroMonthCol).Address($false, $false) + ':' + $cuadro.Cells.Item(7, $cuadroMonthCol).Address($false, $false) + ')'
    $cuadro.Cells.Item(9, $cuadroMonthCol).Formula = '=' + $cuadro.Cells.Item(8, $cuadroMonthCol).Address($false, $false)

    $monthDateSerial = $resumen.Cells.Item(4, $currentStartCol).Value2
    $monthDate = [DateTime]::FromOADate([double]$monthDateSerial)
    $cuadro.Cells.Item(4, $cuadroMonthCol).Value2 = $monthDate.ToOADate()
    $cuadro.Cells.Item(4, $cuadroMonthCol).NumberFormat = 'mmm-yy'
    $daysInMonth = [DateTime]::DaysInMonth($monthDate.Year, $monthDate.Month)
    $cuadro.Cells.Item(12, $cuadroDailyCol).Value2 = $monthDate.ToOADate()
    $cuadro.Cells.Item(12, $cuadroDailyCol).NumberFormat = 'mmm-yy'
    for ($r = 13; $r -le 18; $r++) {{
      $service = [string]$cuadro.Cells.Item($r, 1).Text
      $found = $resumen.Columns.Item(1).Find($service)
      if ($null -ne $found) {{
        $serviceCurrentTotal = [double]$resumen.Cells.Item($found.Row, $currentTotalCol).Value2
        $avgValue = $serviceCurrentTotal / $daysInMonth
        $avgText = [string]::Format([System.Globalization.CultureInfo]::InvariantCulture, '{{0}}', $avgValue)
        $cuadro.Cells.Item($r, $cuadroDailyCol).Formula = '=' + $avgText
      }} else {{
        $cuadro.Cells.Item($r, $cuadroDailyCol).Formula = '=0'
      }}
    }}

    for ($c = 3; $c -le $cuadroMonthCol; $c++) {{
      $cuadro.Cells.Item(8, $c).Formula = '=SUM(' + $cuadro.Cells.Item(5, $c).Address($false, $false) + ':' + $cuadro.Cells.Item(7, $c).Address($false, $false) + ')'
      $cuadro.Cells.Item(9, $c).Formula = '=' + $cuadro.Cells.Item(8, $c).Address($false, $false)
    }}
    if ($cuadro.ChartObjects().Count -gt 0) {{
      $monthNameMap = @{{1='Ene';2='Feb';3='Mar';4='Abr';5='May';6='Jun';7='Jul';8='Ago';9='Sep';10='Oct';11='Nov';12='Dic'}}
      $firstMonthDate = [DateTime]::FromOADate([double]$cuadro.Cells.Item(4, 3).Value2)
      $firstTitleMonth = $monthNameMap[[int]$firstMonthDate.Month] + ' ' + [string]$firstMonthDate.Year
      $lastTitleMonth = $monthNameMap[[int]$monthDate.Month] + ' ' + [string]$monthDate.Year
      for ($chartIndex = 1; $chartIndex -le $cuadro.ChartObjects().Count; $chartIndex++) {{
        $chart = $cuadro.ChartObjects($chartIndex).Chart
        $chartTitle = ''
        try {{ if ($chart.HasTitle) {{ $chartTitle = [string]$chart.ChartTitle.Text }} }} catch {{}}
        if ($chartTitle -like '*Consumos por Ambiente*') {{
          $rangeAddress = $cuadro.Range($cuadro.Cells.Item(4, 2), $cuadro.Cells.Item(7, $cuadroMonthCol))
          $chart.SetSourceData($rangeAddress)
        }} elseif ($chartTitle -like '*Consumos totales por mes*') {{
          $categoriesRange = $cuadro.Range($cuadro.Cells.Item(4, 3), $cuadro.Cells.Item(4, $cuadroMonthCol))
          $totalsRange = $cuadro.Range($cuadro.Cells.Item(8, 3), $cuadro.Cells.Item(8, $cuadroMonthCol))
          if ($chart.SeriesCollection().Count -eq 0) {{ $chart.SeriesCollection().NewSeries() | Out-Null }}
          $series = $chart.SeriesCollection(1)
          $series.XValues = $categoriesRange
          $series.Values = $totalsRange
          $series.Name = 'TOTAL:'
          $chart.HasTitle = $true
          $chart.ChartTitle.Text = 'Consumos totales por mes - IBM Cloud (' + $firstTitleMonth + ' - ' + $lastTitleMonth + ')'
        }} elseif ($chartIndex -eq 1) {{
          $rangeAddress = $cuadro.Range($cuadro.Cells.Item(4, 2), $cuadro.Cells.Item(7, $cuadroMonthCol))
          $chart.SetSourceData($rangeAddress)
        }} elseif ($chartIndex -eq 2) {{
          $categoriesRange = $cuadro.Range($cuadro.Cells.Item(4, 3), $cuadro.Cells.Item(4, $cuadroMonthCol))
          $totalsRange = $cuadro.Range($cuadro.Cells.Item(8, 3), $cuadro.Cells.Item(8, $cuadroMonthCol))
          if ($chart.SeriesCollection().Count -eq 0) {{ $chart.SeriesCollection().NewSeries() | Out-Null }}
          $series = $chart.SeriesCollection(1)
          $series.XValues = $categoriesRange
          $series.Values = $totalsRange
          $series.Name = 'TOTAL:'
          $chart.HasTitle = $true
          $chart.ChartTitle.Text = 'Consumos totales por mes - IBM Cloud (' + $firstTitleMonth + ' - ' + $lastTitleMonth + ')'
        }}
      }}
    }}
    $topRange = $cuadro.Range($cuadro.Cells.Item(4, 2), $cuadro.Cells.Item(8, $cuadroMonthCol))
    $topRange.Borders.LineStyle = 1
    $topRange.Borders.Weight = 2
    $cuadro.Range($cuadro.Cells.Item(4, 2), $cuadro.Cells.Item(4, $cuadroMonthCol)).Interior.Color = $headerColor
    $cuadro.Range($cuadro.Cells.Item(8, 2), $cuadro.Cells.Item(8, $cuadroMonthCol)).Interior.Color = $totalColor
    $cuadro.Range($cuadro.Cells.Item(4, 2), $cuadro.Cells.Item(8, 2)).Font.Bold = $true
    for ($r = 5; $r -le 9; $r++) {{
      for ($c = 3; $c -le $cuadroMonthCol; $c++) {{
        try {{ $cuadro.Cells.Item($r, $c).NumberFormat = $currencyFormat }} catch {{}}
      }}
    }}
    for ($c = 3; $c -le $cuadroMonthCol; $c++) {{
      if ($cuadro.Cells.Item(4, $c).Value2 -is [double] -or $cuadro.Cells.Item(4, $c).Value2 -is [int]) {{
        $cuadro.Cells.Item(4, $c).NumberFormat = 'mmm-yy'
      }} else {{
        $cuadro.Cells.Item(4, $c).NumberFormat = '@'
      }}
      if ($c - 1 -le $cuadroDailyCol -and ($cuadro.Cells.Item(12, $c - 1).Value2 -is [double] -or $cuadro.Cells.Item(12, $c - 1).Value2 -is [int])) {{
        $cuadro.Cells.Item(12, $c - 1).NumberFormat = 'mmm-yy'
      }}
    }}
    $blackRange = $cuadro.Range($cuadro.Cells.Item(12, 1), $cuadro.Cells.Item(18, $cuadroDailyCol))
    $blackRange.Interior.Color = 0
    $blackRange.Font.Color = 16777215
    $blackRange.Borders.LineStyle = 1
    for ($r = 13; $r -le 18; $r++) {{
      for ($c = 2; $c -le $cuadroDailyCol; $c++) {{
        try {{ $cuadro.Cells.Item($r, $c).NumberFormat = $currencyFormat }} catch {{}}
      }}
    }}
    $cuadro.Columns.Item(1).ColumnWidth = 52
    $cuadro.Columns.Item(2).ColumnWidth = 14
    for ($c = 3; $c -le $cuadroMonthCol; $c++) {{ $cuadro.Columns.Item($c).ColumnWidth = 14 }}
  }}
  $resumen.UsedRange.Columns.AutoFit() | Out-Null
  $resumen.Columns.Item(1).ColumnWidth = 52
  for ($c = 2; $c -le $lastCol; $c++) {{
    if ($resumen.Columns.Item($c).ColumnWidth -lt 12) {{ $resumen.Columns.Item($c).ColumnWidth = 12 }}
    if ($resumen.Columns.Item($c).ColumnWidth -gt 18) {{ $resumen.Columns.Item($c).ColumnWidth = 18 }}
  }}
  $resumen.Range($resumen.Cells.Item(1, 1), $resumen.Cells.Item(5, $lastCol)).Interior.ColorIndex = -4142
  if ($totalGeneralCol -gt 0) {{
    $resumen.Range($resumen.Cells.Item(3, 1), $resumen.Cells.Item(5, $totalGeneralCol)).Interior.Color = $headerColor
  }}
  if ($totalVarCol -gt 0) {{
    $variationHeaderStartCol = $totalVarCol - 6
    $resumen.Range($resumen.Cells.Item(3, $variationHeaderStartCol), $resumen.Cells.Item(5, $totalVarCol)).Interior.Color = $headerColor
  }}
  for ($c = 1; $c -le $lastCol; $c++) {{
    $text4 = [string]$resumen.Cells.Item(4, $c).Text
    if ($text4.StartsWith('Total ') -and $text4 -ne 'Total general' -and $text4 -ne 'Total Var.') {{
      $resumen.Range($resumen.Cells.Item(3, $c), $resumen.Cells.Item(55, $c)).Interior.Color = $totalColor
    }} elseif ($text4 -eq 'Total general' -or $text4 -eq 'Total Var.') {{
      $resumen.Range($resumen.Cells.Item(3, $c), $resumen.Cells.Item(5, $c)).Interior.Color = $totalColor
    }}
  }}
  if ($totalGeneralCol -gt 0) {{
    $resumen.Range($resumen.Cells.Item(6, $totalGeneralCol), $resumen.Cells.Item($lastRow, $totalGeneralCol)).Interior.ColorIndex = -4142
  }}
  if ($totalVarCol -gt 0) {{
    $resumen.Range($resumen.Cells.Item(6, $totalVarCol), $resumen.Cells.Item($lastRow, $totalVarCol)).Interior.ColorIndex = -4142
    for ($r = 6; $r -le $totalGeneralRow; $r++) {{
      $value = $resumen.Cells.Item($r, $totalVarCol).Value2
        if ($value -is [double] -or $value -is [int]) {{
          if ([math]::Abs([double]$value) -lt 0.005) {{
            $resumen.Cells.Item($r, $totalVarCol).Value2 = '-'
            $resumen.Cells.Item($r, $totalVarCol).Interior.ColorIndex = -4142
          }} elseif ($value -gt 0) {{
            $resumen.Cells.Item($r, $totalVarCol).Interior.Color = 255
          }} elseif ($value -lt 0) {{
            $resumen.Cells.Item($r, $totalVarCol).Interior.Color = 5296274
        }}
      }}
    }}
  }}
  if ($totalVarCol -gt 0 -and $totalGeneralRow -gt 0) {{
    $variationStartCol = $totalVarCol - 6
    $variationDataRange = $resumen.Range($resumen.Cells.Item(6, $variationStartCol), $resumen.Cells.Item($totalGeneralRow + 1, $totalVarCol))
    try {{ $variationDataRange.FormatConditions.Delete() | Out-Null }} catch {{}}
    for ($r = 6; $r -le ($totalGeneralRow + 1); $r++) {{
      for ($c = $variationStartCol; $c -le $totalVarCol; $c++) {{
        $cell = $resumen.Cells.Item($r, $c)
        $textValue = ([string]$cell.Text).Trim()
        if ($textValue -eq '-') {{
          $cell.Interior.ColorIndex = -4142
        }}
      }}
    }}
  }}
  $resumen.Range($resumen.Cells.Item(3, 1), $resumen.Cells.Item(5, $lastCol)).Font.Bold = $true
  $resumen.Range($resumen.Cells.Item(6, 1), $resumen.Cells.Item($totalGeneralRow, $lastCol)).Font.Bold = $false
  for ($r = 6; $r -le $totalGeneralRow; $r++) {{
    $label = ([string]$resumen.Cells.Item($r, 1).Text).Trim()
    if ($label -eq '') {{ continue }}
    $indent = 0
    $nextIndent = 0
    try {{ $indent = [int]$resumen.Cells.Item($r, 1).IndentLevel }} catch {{}}
    try {{ $nextIndent = [int]$resumen.Cells.Item($r + 1, 1).IndentLevel }} catch {{}}
    $isServiceRow = ($indent -eq 0 -and $nextIndent -gt $indent)
    if ($isServiceRow -or $label -eq 'Total general') {{
      $resumen.Range($resumen.Cells.Item($r, 1), $resumen.Cells.Item($r, $lastCol)).Font.Bold = $true
    }}
  }}
  if ($totalGeneralCol -gt 0 -and $resumen.Columns.Item($totalGeneralCol).ColumnWidth -lt 16) {{
    $resumen.Columns.Item($totalGeneralCol).ColumnWidth = 16
  }}
  $resumen.Range($resumen.Cells.Item(3, 1), $resumen.Cells.Item($lastRow, $lastCol)).Borders.LineStyle = 1
  foreach ($sheetName in @('Variacion Barel Metal', 'VS&AS')) {{
    try {{
      $wsVar = $wb.Worksheets.Item($sheetName)
      $wsVar.Cells.Replace('(en blanco)', '', 1, 1, $false, $false, $false, $false) | Out-Null
      $wsVar.UsedRange.Columns.AutoFit() | Out-Null
      $wsVar.Columns.Item(1).ColumnWidth = 46
      $varLastCol = $wsVar.UsedRange.Column + $wsVar.UsedRange.Columns.Count - 1
      for ($c = 2; $c -le $varLastCol; $c++) {{
        if ($wsVar.Columns.Item($c).ColumnWidth -lt 12) {{ $wsVar.Columns.Item($c).ColumnWidth = 12 }}
        if ($wsVar.Columns.Item($c).ColumnWidth -gt 18) {{ $wsVar.Columns.Item($c).ColumnWidth = 18 }}
      }}
    }} catch {{}}
  }}
  $excel.CalculateFull()
  $wb.SaveAs('{str(temporal).replace("'", "''")}', 51)
  $wb.Close($false)
}} finally {{
  $excel.Quit()
  [System.Runtime.InteropServices.Marshal]::ReleaseComObject($excel) | Out-Null
}}
"""
    try:
        subprocess.run(
            ["powershell", "-NoProfile", "-ExecutionPolicy", "Bypass", "-Command", ps_script],
            check=True,
            capture_output=True,
            text=True,
            timeout=90,
        )
        temporal.replace(ruta)
        return True
    except subprocess.TimeoutExpired:
        print("Advertencia: Excel no termino el post-proceso dentro del tiempo limite.")
        if temporal.exists():
            temporal.unlink()
        return False
    except subprocess.CalledProcessError as exc:
        detalle = (exc.stderr or exc.stdout or str(exc)).strip()
        if detalle:
            print(f"Advertencia: Excel no pudo aplicar el post-proceso: {detalle}")
        if temporal.exists():
            temporal.unlink()
        return False
    except Exception as exc:
        print(f"Advertencia: Excel no pudo aplicar el post-proceso: {exc}")
        if temporal.exists():
            temporal.unlink()
        return False
    finally:
        if marcas_amarillas_path.exists():
            marcas_amarillas_path.unlink()


def normalizar_entorno(valor: Any) -> str:
    if valor is None:
        return ""
    valor = str(valor).strip()
    if valor.lower() in {"", "none", "nan"}:
        return ""
    return valor


def a_numero(valor: Any) -> float:
    if valor is None or valor == "":
        return 0.0
    try:
        return float(valor)
    except Exception:
        return 0.0


def leer_filas_excel(ruta: str | Path) -> list[dict[str, Any]]:
    wb = load_workbook(ruta, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    filas = list(ws.iter_rows(values_only=True))
    wb.close()
    if not filas:
        return []
    encabezados = [str(c).strip() if c is not None else "" for c in filas[0]]
    salida: list[dict[str, Any]] = []
    for fila in filas[1:]:
        salida.append({encabezados[i]: fila[i] if i < len(fila) else None for i in range(len(encabezados))})
    return salida


def consolidar_summary(ruta_summary: str | Path):
    filas = leer_filas_excel(ruta_summary)
    resumen = defaultdict(float)
    for f in filas:
        servicio = f.get("Service Name")
        plan = f.get("Plan Name")
        if servicio and plan:
            resumen[(str(servicio), str(plan))] += a_numero(f.get("Cost"))
    return resumen


def consolidar_instances(ruta_instances: str | Path):
    filas = leer_filas_excel(ruta_instances)
    por_plan = defaultdict(lambda: defaultdict(float))
    por_servicio = defaultdict(lambda: defaultdict(float))
    servicios_planes = defaultdict(set)
    for f in filas:
        servicio = f.get("Service Name")
        plan = f.get("Plan Name")
        if not servicio or not plan:
            continue
        servicio = str(servicio)
        plan = str(plan)
        entorno = normalizar_entorno(f.get("entorno"))
        costo = a_numero(f.get("Cost"))
        por_plan[(servicio, plan)][entorno] += costo
        por_servicio[servicio][entorno] += costo
        servicios_planes[servicio].add(plan)
    return por_plan, por_servicio, servicios_planes


def aplicar_formato_tabla(ws, max_row: int, max_col: int):
    fill_header = PatternFill("solid", fgColor="D9E1F2")
    fill_total = PatternFill("solid", fgColor="D9D9D9")
    thin_blue = Side(style="thin", color="45B6E8")
    medium_blue = Side(style="medium", color="45B6E8")
    border_row = Border()
    border_header = Border(bottom=thin_blue)
    border_service = Border(bottom=medium_blue)
    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.border = border_row
            cell.alignment = Alignment(
                horizontal=cell.alignment.horizontal,
                vertical="center",
                indent=cell.alignment.indent,
            )
            if isinstance(cell.value, (int, float)):
                if abs(float(cell.value)) < 0.0000001:
                    cell.value = None
                else:
                    cell.number_format = '$#,##0.00;[Red]-$#,##0.00;'
    for cell in ws[1]:
        cell.fill = fill_header
        cell.font = Font(bold=True)
        cell.border = border_header
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row_idx in range(2, max_row + 1):
        first = ws.cell(row_idx, 1).value
        if first == "Total general":
            for cell in ws[row_idx]:
                cell.fill = fill_total
                cell.font = Font(bold=True)
                cell.border = border_header
        elif first and ws.cell(row_idx, 1).alignment.indent == 0 and row_idx < max_row and ws.cell(row_idx + 1, 1).alignment.indent > 0:
            for cell in ws[row_idx]:
                cell.fill = PatternFill(fill_type=None)
                cell.font = Font(bold=True)
                cell.border = border_service
        elif first:
            ws.cell(row_idx, 1).alignment = Alignment(vertical="center", indent=ws.cell(row_idx, 1).alignment.indent)
    ws.freeze_panes = "A2"
    for col in range(1, max_col + 1):
        letra = get_column_letter(col)
        ws.column_dimensions[letra].width = 48 if col == 1 else 18
    ws.sheet_view.showGridLines = False


def generar_excels_tablas(ruta_summary: str | Path, ruta_instances: str | Path, summary, por_plan, por_servicio, servicios_planes):
    entornos_instances = ["", "all", "dev", "devqa", "prod", "qa"]

    wb_summary = Workbook()
    ws1 = wb_summary.active
    ws1.title = "XM_Summary"
    ws1.append(["Etiquetas de fila", "Suma de Cost"])
    total_summary = 0.0
    for servicio in sorted({s for s, _ in summary.keys()}):
        total_servicio = sum(v for (s, _), v in summary.items() if s == servicio)
        ws1.append([servicio, total_servicio])
        total_summary += total_servicio
        for _, plan in sorted(k for k in summary if k[0] == servicio):
            ws1.append([plan, summary[(servicio, plan)]])
            ws1.cell(ws1.max_row, 1).alignment = Alignment(indent=1)
    ws1.append(["Total general", total_summary])
    ws1.cell(ws1.max_row, 1).font = Font(bold=True)
    ws1.cell(ws1.max_row, 2).font = Font(bold=True)
    aplicar_formato_tabla(ws1, ws1.max_row, 2)
    wb_summary.save(ruta_summary)
    wb_summary.close()

    wb_instances = Workbook()
    ws2 = wb_instances.active
    ws2.title = "XM_Instances"
    ws2.append(["Etiquetas de fila", "", "all", "dev", "devqa", "prod", "qa", "Total general"])
    for servicio in sorted(por_servicio.keys()):
        d = por_servicio[servicio]
        vals = [d.get(e, 0.0) for e in entornos_instances]
        ws2.append([servicio, *vals, sum(vals)])
        for plan in sorted(servicios_planes[servicio]):
            dp = por_plan[(servicio, plan)]
            vals = [dp.get(e, 0.0) for e in entornos_instances]
            ws2.append([plan, *vals, sum(vals)])
            ws2.cell(ws2.max_row, 1).alignment = Alignment(indent=1)
    totales = [sum(por_servicio[s].get(e, 0.0) for s in por_servicio) for e in entornos_instances]
    ws2.append(["Total general", *totales, sum(totales)])
    for c in range(1, 9):
        ws2.cell(ws2.max_row, c).font = Font(bold=True)
    aplicar_formato_tabla(ws2, ws2.max_row, 8)
    wb_instances.save(ruta_instances)
    wb_instances.close()


def col_to_index(col: str) -> int:
    n = 0
    for ch in col:
        n = n * 26 + ord(ch) - 64
    return n


def index_to_col(n: int) -> str:
    s = ""
    while n:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


def split_ref(ref: str):
    m = re.match(r"([A-Z]+)(\d+)", ref)
    return col_to_index(m.group(1)), int(m.group(2))


def parse_range_ref(ref: str):
    if ":" not in ref:
        col, row = split_ref(ref)
        return col, row, col, row
    start, end = ref.split(":", 1)
    start_col, start_row = split_ref(start)
    end_col, end_row = split_ref(end)
    return start_col, start_row, end_col, end_row


def format_range_ref(start_col: int, start_row: int, end_col: int, end_row: int) -> str:
    start = f"{index_to_col(start_col)}{start_row}"
    end = f"{index_to_col(end_col)}{end_row}"
    return start if start == end else f"{start}:{end}"


def shift_range_ref(ref: str, insert_col: int, amount: int) -> str:
    start_col, start_row, end_col, end_row = parse_range_ref(ref)
    if start_col >= insert_col:
        start_col += amount
    if end_col >= insert_col:
        end_col += amount
    return format_range_ref(start_col, start_row, end_col, end_row)


def parse_etiqueta_mes(etiqueta_mes: str) -> tuple[int, int, str]:
    m = re.fullmatch(r"([A-Za-zÁÉÍÓÚáéíóúñÑ]{3})-(\d{2}|\d{4})", etiqueta_mes.strip())
    if not m:
        raise ValueError("El mes debe tener formato como Jul-26.")
    mes_txt = m.group(1).lower()
    mes_txt = (
        mes_txt.replace("á", "a")
        .replace("é", "e")
        .replace("í", "i")
        .replace("ó", "o")
        .replace("ú", "u")
    )
    if mes_txt not in MESES:
        raise ValueError(f"Mes no reconocido: {etiqueta_mes}.")
    year_raw = int(m.group(2))
    year = 2000 + year_raw if year_raw < 100 else year_raw
    mes = MESES[mes_txt]
    etiqueta_normalizada = f"{MESES_ABREV[mes]}-{str(year)[-2:]}"
    return year, mes, etiqueta_normalizada


def excel_serial(fecha: date) -> int:
    return (fecha - date(1899, 12, 30)).days


def load_shared(z: zipfile.ZipFile):
    try:
        root = ET.fromstring(z.read("xl/sharedStrings.xml"))
    except KeyError:
        return []
    return ["".join(t.text or "" for t in si.iter(NS + "t")) for si in root.findall(NS + "si")]


def cell_value(c, shared):
    if c is None:
        return None
    t = c.attrib.get("t")
    v = c.find(NS + "v")
    isel = c.find(NS + "is")
    if t == "s" and v is not None:
        return shared[int(v.text)]
    if t == "inlineStr" and isel is not None:
        return "".join(tt.text or "" for tt in isel.iter(NS + "t"))
    if v is not None:
        txt = v.text
        try:
            fl = float(txt)
            return int(fl) if fl.is_integer() else fl
        except Exception:
            return txt
    return None


def set_number(c, value):
    for child in list(c):
        if child.tag in (NS + "f", NS + "v", NS + "is"):
            c.remove(child)
    c.attrib.pop("t", None)
    if value is None:
        return
    v = ET.SubElement(c, NS + "v")
    v.text = repr(float(value)) if isinstance(value, float) else str(value)


def set_text(c, value: str):
    for child in list(c):
        if child.tag in (NS + "f", NS + "v", NS + "is"):
            c.remove(child)
    if SHARED_STRINGS_CONTEXT is not None:
        shared, shared_root = SHARED_STRINGS_CONTEXT
        shared_root.attrib["count"] = str(int(shared_root.attrib.get("count", "0")) + 1)
        if value in shared:
            idx = shared.index(value)
        else:
            idx = len(shared)
            shared.append(value)
            si = ET.SubElement(shared_root, NS + "si")
            t = ET.SubElement(si, NS + "t")
            t.text = value
            shared_root.attrib["uniqueCount"] = str(len(shared))
        c.attrib["t"] = "s"
        v = ET.SubElement(c, NS + "v")
        v.text = str(idx)
    else:
        c.attrib["t"] = "inlineStr"
        isel = ET.SubElement(c, NS + "is")
        t = ET.SubElement(isel, NS + "t")
        t.text = value


def set_formula(c, formula: str):
    for child in list(c):
        if child.tag in (NS + "f", NS + "v", NS + "is"):
            c.remove(child)
    c.attrib.pop("t", None)
    f = ET.SubElement(c, NS + "f")
    f.text = formula


def copy_style(source, target):
    if source is not None and "s" in source.attrib:
        target.attrib["s"] = source.attrib["s"]
    elif "s" in target.attrib:
        target.attrib.pop("s", None)


def get_or_create_cell(row_elem, col_idx: int, row_idx: int):
    ref = f"{index_to_col(col_idx)}{row_idx}"
    for c in row_elem.findall(NS + "c"):
        if c.attrib.get("r") == ref:
            return c
    new = ET.Element(NS + "c", {"r": ref})
    children = list(row_elem)
    for c in row_elem.findall(NS + "c"):
        ci, _ = split_ref(c.attrib["r"])
        if ci > col_idx:
            row_elem.insert(children.index(c), new)
            return new
    row_elem.append(new)
    return new


def find_cell(row_elem, col_idx: int, row_idx: int):
    ref = f"{index_to_col(col_idx)}{row_idx}"
    for c in row_elem.findall(NS + "c"):
        if c.attrib.get("r") == ref:
            return c
    return None


def get_or_create_row(sheet_data, row_idx: int):
    for row in sheet_data.findall(NS + "row"):
        if int(row.attrib["r"]) == row_idx:
            return row
    new = ET.Element(NS + "row", {"r": str(row_idx)})
    for row in sheet_data.findall(NS + "row"):
        if int(row.attrib["r"]) > row_idx:
            sheet_data.insert(list(sheet_data).index(row), new)
            return new
    sheet_data.append(new)
    return new


def shift_sheet_columns(root, insert_col: int, amount: int):
    dimension = root.find(NS + "dimension")
    if dimension is not None and "ref" in dimension.attrib:
        dimension.attrib["ref"] = shift_range_ref(dimension.attrib["ref"], insert_col, amount)

    cols = root.find(NS + "cols")
    if cols is not None:
        for col in cols.findall(NS + "col"):
            min_col = int(col.attrib["min"])
            max_col = int(col.attrib["max"])
            if min_col >= insert_col:
                col.attrib["min"] = str(min_col + amount)
                col.attrib["max"] = str(max_col + amount)
            elif max_col >= insert_col:
                col.attrib["max"] = str(max_col + amount)

    sheet_data = root.find(NS + "sheetData")
    for row in sheet_data.findall(NS + "row"):
        if "spans" in row.attrib:
            start, end = row.attrib["spans"].split(":")
            end_i = int(end)
            if end_i >= insert_col:
                row.attrib["spans"] = f"{start}:{end_i + amount}"
        for cell in row.findall(NS + "c"):
            ref = cell.attrib.get("r")
            if not ref:
                continue
            col_idx, row_idx = split_ref(ref)
            if col_idx >= insert_col:
                cell.attrib["r"] = f"{index_to_col(col_idx + amount)}{row_idx}"

    merge_cells = root.find(NS + "mergeCells")
    if merge_cells is not None:
        for merge_cell in merge_cells.findall(NS + "mergeCell"):
            merge_cell.attrib["ref"] = shift_range_ref(merge_cell.attrib["ref"], insert_col, amount)

    for conditional in root.findall(NS + "conditionalFormatting"):
        refs = conditional.attrib.get("sqref", "").split()
        conditional.attrib["sqref"] = " ".join(shift_range_ref(ref, insert_col, amount) for ref in refs)


def namespace_is_used(root, namespace_uri: str) -> bool:
    marker = "{" + namespace_uri + "}"
    for elem in root.iter():
        if isinstance(elem.tag, str) and elem.tag.startswith(marker):
            return True
        if any(isinstance(attr, str) and attr.startswith(marker) for attr in elem.attrib):
            return True
    return False


def normalize_ignorable_prefixes(root):
    attr = "{" + NS_MC + "}Ignorable"
    value = root.attrib.get(attr)
    if not value:
        return
    kept = [
        prefix
        for prefix in value.split()
        if prefix in KNOWN_PREFIX_NAMESPACES and namespace_is_used(root, KNOWN_PREFIX_NAMESPACES[prefix])
    ]
    if kept:
        root.attrib[attr] = " ".join(kept)
    else:
        root.attrib.pop(attr, None)


def actualizar_cuadro_xml(root, shared, etiqueta_mes: str, year: int, mes: int, por_servicio):
    sheet_data = root.find(NS + "sheetData")
    rows = {int(r.attrib["r"]): r for r in sheet_data.findall(NS + "row")}
    days = calendar.monthrange(year, mes)[1]

    month_cols = []
    row4 = rows.get(4)
    if row4 is not None:
        for cell in row4.findall(NS + "c"):
            value = str(cell_value(cell, shared)).strip()
            if re.fullmatch(r"[A-Z][a-z]{2}-\d{2}", value):
                month_cols.append(split_ref(cell.attrib["r"])[0])
    if not month_cols:
        return

    target_col = None
    for col in month_cols:
        if str(cell_value(find_cell(row4, col, 4), shared)).strip() == etiqueta_mes:
            target_col = col
            break
    previous_col = max(month_cols)
    if target_col is None:
        target_col = previous_col + 1

    for row_idx in range(4, 19):
        row = get_or_create_row(sheet_data, row_idx)
        source = find_cell(row, previous_col, row_idx)
        target = get_or_create_cell(row, target_col, row_idx)
        copy_style(source, target)

    set_text(get_or_create_cell(rows[4], target_col, 4), etiqueta_mes)
    ambiente_totales = {
        "dev": sum(v.get("dev", 0.0) for v in por_servicio.values()),
        "qa": sum(v.get("qa", 0.0) for v in por_servicio.values()),
        "prod": sum(v.get("prod", 0.0) for v in por_servicio.values()),
    }
    set_number(get_or_create_cell(rows[5], target_col, 5), ambiente_totales["dev"])
    set_number(get_or_create_cell(rows[6], target_col, 6), ambiente_totales["qa"])
    set_number(get_or_create_cell(rows[7], target_col, 7), ambiente_totales["prod"])
    set_number(get_or_create_cell(rows[8], target_col, 8), sum(ambiente_totales.values()))
    if 9 in rows:
        set_number(get_or_create_cell(rows[9], target_col, 9), sum(ambiente_totales.values()))

    row12 = get_or_create_row(sheet_data, 12)
    set_text(get_or_create_cell(row12, target_col, 12), etiqueta_mes)
    for row_idx in range(13, 19):
        row = rows.get(row_idx)
        if row is None:
            continue
        service = cell_value(find_cell(row, 1, row_idx), shared)
        if service is None:
            continue
        service = str(service).strip()
        monthly_total = sum(por_servicio.get(service, {}).values())
        set_number(get_or_create_cell(row, target_col, row_idx), monthly_total / days if monthly_total else 0.0)


def actualizar_comparativo_xml(
    ruta_base: str | Path,
    ruta_salida: str | Path,
    etiqueta_mes: str,
    por_plan,
    por_servicio,
    accion_mes: str = "auto",
    ruta_instances: str | Path | None = None,
    postproceso_excel: bool = True,
):
    with zipfile.ZipFile(ruta_base, "r") as zin:
        shared = load_shared(zin)
        ns = {"m": NS_MAIN, "r": NS_R}
        wb = ET.fromstring(zin.read("xl/workbook.xml"))
        rels = ET.fromstring(zin.read("xl/_rels/workbook.xml.rels"))
        relmap = {rel.attrib["Id"]: rel.attrib["Target"] for rel in rels}
        resumen_path = None
        for s in wb.find("m:sheets", ns):
            if s.attrib["name"] == "Resumen":
                resumen_path = "xl/" + relmap[s.attrib["{%s}id" % NS_R]].lstrip("/")
                break
        if not resumen_path:
            raise ValueError("No se encontró la hoja Resumen.")

        root = ET.fromstring(zin.read(resumen_path))
        sheet_data = root.find(NS + "sheetData")
        rows = {int(r.attrib["r"]): r for r in sheet_data.findall(NS + "row")}

        year, mes, etiqueta_mes = parse_etiqueta_mes(etiqueta_mes)
        total_col = None
        total_general_col = None
        month_total_cols = []
        # En este archivo, los encabezados de mes están en fila 4.
        for c in rows.get(4, []).findall(NS + "c"):
            valor = str(cell_value(c, shared)).strip()
            col_idx = split_ref(c.attrib["r"])[0]
            if valor == f"Total {etiqueta_mes}":
                total_col = col_idx
            elif valor == "Total general":
                total_general_col = col_idx
            elif valor.startswith("Total ") and valor not in {"Total general", "Total Var."}:
                month_total_cols.append(col_idx)

        if not total_general_col:
            raise ValueError("No se encontro la columna 'Total general' en la fila 4 del comparativo.")

        if total_col and accion_mes == "agregar":
            total_col = None

        if total_col:
            start_col = total_col - 6
            previous_start_col = start_col - 7 if start_col > 8 else start_col
        else:
            if not month_total_cols:
                raise ValueError("No se encontraron bloques de meses en la fila 4 del comparativo.")
            previous_total_col = max(col for col in month_total_cols if col < total_general_col)
            previous_start_col = previous_total_col - 6
            start_col = total_general_col
            total_col = start_col + 6
            shift_sheet_columns(root, total_general_col, 7)
            total_general_col += 7
            month_total_cols.append(total_col)

            rows = {int(r.attrib["r"]): r for r in sheet_data.findall(NS + "row")}
            row4 = rows[4]
            row5 = rows[5]
            for i in range(7):
                source4 = find_cell(row4, previous_start_col + i, 4)
                target4 = get_or_create_cell(row4, start_col + i, 4)
                copy_style(source4, target4)
                source5 = find_cell(row5, previous_start_col + i, 5)
                target5 = get_or_create_cell(row5, start_col + i, 5)
                copy_style(source5, target5)

            set_number(get_or_create_cell(row4, start_col, 4), excel_serial(date(year, mes, 1)))
            for i in range(1, 6):
                set_number(get_or_create_cell(row4, start_col + i, 4), None)
            set_text(get_or_create_cell(row4, total_col, 4), f"Total {etiqueta_mes}")

            for i, label in enumerate(["dev", "devqa", "qa", "", "prod", "all"]):
                set_text(get_or_create_cell(row5, start_col + i, 5), label)
            set_number(get_or_create_cell(row5, total_col, 5), None)

        month_total_cols = sorted(set(month_total_cols))
        entornos = ["dev", "devqa", "qa", "", "prod", "all"]
        variation_start_col = total_general_col + 3
        previous_label = str(cell_value(find_cell(rows[4], start_col - 1, 4), shared) or "").replace("Total ", "")
        variation_label = f"Variacion {previous_label}-{etiqueta_mes}"
        servicio_actual = None

        for r_idx in range(6, max(rows) + 1):
            row = rows.get(r_idx)
            if row is None:
                continue
            for i in range(7):
                source = find_cell(row, previous_start_col + i, r_idx)
                target = get_or_create_cell(row, start_col + i, r_idx)
                copy_style(source, target)
            a_cell = None
            for c in row.findall(NS + "c"):
                if c.attrib.get("r") == f"A{r_idx}":
                    a_cell = c
                    break
            etiqueta = cell_value(a_cell, shared)
            if etiqueta is None:
                continue
            etiqueta = str(etiqueta).strip()
            datos = None
            if etiqueta == "Total general":
                datos = {e: sum(por_servicio[s].get(e, 0.0) for s in por_servicio) for e in entornos}
            elif etiqueta in por_servicio:
                servicio_actual = etiqueta
                datos = por_servicio[etiqueta]
            elif servicio_actual and (servicio_actual, etiqueta) in por_plan:
                datos = por_plan[(servicio_actual, etiqueta)]

            if datos is None:
                continue

            total = 0.0
            for i, entorno in enumerate(entornos):
                valor = float(datos.get(entorno, 0.0))
                total += valor
                set_number(get_or_create_cell(row, start_col + i, r_idx), valor if (valor != 0 or etiqueta == "Total general") else None)
            set_number(get_or_create_cell(row, total_col, r_idx), total)
            total_general = sum(a_numero(cell_value(find_cell(row, col, r_idx), shared)) for col in month_total_cols)
            set_number(get_or_create_cell(row, total_general_col, r_idx), total_general)

        row4 = rows[4]
        row5 = rows[5]
        set_text(get_or_create_cell(row4, variation_start_col, 4), variation_label)
        set_text(get_or_create_cell(row4, variation_start_col + 6, 4), "Total Var.")
        for i, label in enumerate(["dev", "devqa", "qa", "", "prod", "all"]):
            set_text(get_or_create_cell(row5, variation_start_col + i, 5), label)
        set_number(get_or_create_cell(row5, variation_start_col + 6, 5), None)

        for r_idx in range(6, max(rows) + 1):
            row = rows.get(r_idx)
            if row is None or find_cell(row, 1, r_idx) is None:
                continue
            for i in range(7):
                new_ref = f"{index_to_col(start_col + i)}{r_idx}"
                old_ref = f"{index_to_col(previous_start_col + i)}{r_idx}"
                target = get_or_create_cell(row, variation_start_col + i, r_idx)
                set_formula(
                    target,
                    f'IF(ABS(N({new_ref})-N({old_ref}))<0.005,"-",N({new_ref})-N({old_ref}))',
                )

        with zipfile.ZipFile(ruta_salida, "w", zipfile.ZIP_DEFLATED) as zout:
            rels_calc = [
                rel
                for rel in list(rels)
                if rel.attrib.get("Type", "").endswith("/calcChain")
            ]
            for rel in rels_calc:
                rels.remove(rel)
            content_types = ET.fromstring(zin.read("[Content_Types].xml"))
            for override in list(content_types):
                if override.attrib.get("PartName") == "/xl/calcChain.xml":
                    content_types.remove(override)
            for item in zin.infolist():
                if item.filename == "xl/calcChain.xml":
                    continue
                data = zin.read(item.filename)
                if item.filename == resumen_path:
                    data = ET.tostring(root, encoding="utf-8", xml_declaration=True)
                elif item.filename == "xl/_rels/workbook.xml.rels":
                    data = ET.tostring(rels, encoding="utf-8", xml_declaration=True)
                elif item.filename == "[Content_Types].xml":
                    data = ET.tostring(content_types, encoding="utf-8", xml_declaration=True)
                zout.writestr(item, data)


def actualizar_comparativo_xml(
    ruta_base: str | Path,
    ruta_salida: str | Path,
    etiqueta_mes: str,
    por_plan,
    por_servicio,
    accion_mes: str = "auto",
    ruta_instances: str | Path | None = None,
    postproceso_excel: bool = True,
):
    year, mes, etiqueta_mes = parse_etiqueta_mes(etiqueta_mes)
    wb = load_workbook(ruta_base)
    if "Resumen" not in wb.sheetnames:
        wb.close()
        raise ValueError("No se encontro la hoja Resumen.")

    ws = wb["Resumen"]
    total_col = None
    total_general_col = None
    month_total_cols: list[int] = []

    for cell in ws[4]:
        value = str(cell.value).strip() if cell.value is not None else ""
        if value == f"Total {etiqueta_mes}":
            total_col = cell.column
        elif value == "Total general":
            total_general_col = cell.column
        elif value.startswith("Total ") and value not in {"Total general", "Total Var."}:
            month_total_cols.append(cell.column)

    if not total_general_col:
        wb.close()
        raise ValueError("No se encontro la columna 'Total general' en la fila 4 del comparativo.")

    if total_col and accion_mes == "agregar":
        total_col = None

    if total_col:
        start_col = total_col - 6
        previous_start_col = start_col - 7 if start_col > 8 else start_col
    else:
        if not month_total_cols:
            wb.close()
            raise ValueError("No se encontraron bloques de meses en la fila 4 del comparativo.")

        previous_total_col = max(col for col in month_total_cols if col < total_general_col)
        previous_start_col = previous_total_col - 6
        start_col = total_general_col
        ws.insert_cols(start_col, 7)
        total_col = start_col + 6
        total_general_col += 7
        month_total_cols.append(total_col)

        for offset in range(7):
            source_letter = get_column_letter(previous_start_col + offset)
            target_letter = get_column_letter(start_col + offset)
            ws.column_dimensions[target_letter].width = ws.column_dimensions[source_letter].width

        for row_idx in range(1, ws.max_row + 1):
            for offset in range(7):
                source = ws.cell(row_idx, previous_start_col + offset)
                target = ws.cell(row_idx, start_col + offset)
                if source.has_style:
                    target._style = copy.copy(source._style)
                if source.number_format:
                    target.number_format = source.number_format
                if source.alignment:
                    target.alignment = copy.copy(source.alignment)

        for merged_range in list(ws.merged_cells.ranges):
            if merged_range.min_row == 4 and merged_range.max_row == 4 and merged_range.min_col >= total_general_col:
                ws.unmerge_cells(str(merged_range))

        ws.cell(4, start_col).value = date(year, mes, 1)
        for offset in range(1, 6):
            ws.cell(4, start_col + offset).value = None
        ws.cell(4, total_col).value = f"Total {etiqueta_mes}"

        for offset, label in enumerate(["dev", "devqa", "qa", "", "prod", "all"]):
            ws.cell(5, start_col + offset).value = label
        ws.cell(5, total_col).value = None

    month_total_cols = sorted(set(month_total_cols))
    entornos = ["dev", "devqa", "qa", "", "prod", "all"]
    variation_start_col = total_general_col + 3
    previous_label = str(ws.cell(4, start_col - 1).value or "").replace("Total ", "")
    ws.cell(4, variation_start_col).value = f"Variacion {previous_label}-{etiqueta_mes}"
    ws.cell(4, variation_start_col + 6).value = "Total Var."
    for offset, label in enumerate(["dev", "devqa", "qa", "", "prod", "all"]):
        ws.cell(5, variation_start_col + offset).value = label
    ws.cell(5, variation_start_col + 6).value = None

    for merged_range in list(ws.merged_cells.ranges):
        if merged_range.min_row == 4 and merged_range.max_row == 4 and merged_range.min_col >= variation_start_col:
            ws.unmerge_cells(str(merged_range))
    ws.merge_cells(
        start_row=4,
        start_column=variation_start_col,
        end_row=4,
        end_column=variation_start_col + 5,
    )

    servicio_actual = None
    for row_idx in range(6, ws.max_row + 1):
        etiqueta = ws.cell(row_idx, 1).value
        if etiqueta is None:
            continue
        etiqueta = str(etiqueta).strip()

        datos = None
        if etiqueta == "Total general":
            datos = {e: sum(por_servicio[s].get(e, 0.0) for s in por_servicio) for e in entornos}
        elif etiqueta in por_servicio:
            servicio_actual = etiqueta
            datos = por_servicio[etiqueta]
        elif servicio_actual and (servicio_actual, etiqueta) in por_plan:
            datos = por_plan[(servicio_actual, etiqueta)]

        if datos is not None:
            total = 0.0
            for offset, entorno in enumerate(entornos):
                value = float(datos.get(entorno, 0.0))
                total += value
                ws.cell(row_idx, start_col + offset).value = value if (value != 0 or etiqueta == "Total general") else None
            ws.cell(row_idx, total_col).value = total
            ws.cell(row_idx, total_general_col).value = sum(
                a_numero(ws.cell(row_idx, col).value) for col in month_total_cols
            )

        for offset in range(7):
            new_ref = f"{get_column_letter(start_col + offset)}{row_idx}"
            old_ref = f"{get_column_letter(previous_start_col + offset)}{row_idx}"
            ws.cell(row_idx, variation_start_col + offset).value = (
                f'=IF(ABS(N({new_ref})-N({old_ref}))<0.005,"-",N({new_ref})-N({old_ref}))'
            )

    wb.save(ruta_salida)
    wb.close()
    if postproceso_excel and not reparar_con_excel_si_disponible(ruta_salida, ruta_instances, etiqueta_mes, accion_mes):
        raise RuntimeError("Excel no pudo actualizar el comparativo. Cierra cualquier ventana de Excel abierta y vuelve a intentar.")


def actualizar_comparativo_xml(
    ruta_base: str | Path,
    ruta_salida: str | Path,
    etiqueta_mes: str,
    por_plan,
    por_servicio,
    accion_mes: str = "auto",
    ruta_instances: str | Path | None = None,
    postproceso_excel: bool = True,
):
    global SHARED_STRINGS_CONTEXT
    with zipfile.ZipFile(ruta_base, "r") as zin:
        shared = load_shared(zin)
        try:
            shared_root = ET.fromstring(zin.read("xl/sharedStrings.xml"))
        except KeyError:
            shared_root = ET.Element(NS + "sst", {"count": "0", "uniqueCount": "0"})
        SHARED_STRINGS_CONTEXT = (shared, shared_root)

        ns = {"m": NS_MAIN, "r": NS_R}
        wb = ET.fromstring(zin.read("xl/workbook.xml"))
        rels = ET.fromstring(zin.read("xl/_rels/workbook.xml.rels"))
        relmap = {rel.attrib["Id"]: rel.attrib["Target"] for rel in rels}
        resumen_path = None
        cuadro_path = None
        for s in wb.find("m:sheets", ns):
            if s.attrib["name"] == "Resumen":
                resumen_path = "xl/" + relmap[s.attrib["{%s}id" % NS_R]].lstrip("/")
            elif s.attrib["name"] == "Cuadro":
                cuadro_path = "xl/" + relmap[s.attrib["{%s}id" % NS_R]].lstrip("/")
        if not resumen_path:
            SHARED_STRINGS_CONTEXT = None
            raise ValueError("No se encontro la hoja Resumen.")

        resumen_rels_path = resumen_path.replace("worksheets/", "worksheets/_rels/") + ".rels"
        removed_resumen_relationship_parts = set()
        resumen_rels = None
        try:
            resumen_rels = ET.fromstring(zin.read(resumen_rels_path))
        except KeyError:
            resumen_rels = None

        root = ET.fromstring(zin.read(resumen_path))
        cuadro_root = ET.fromstring(zin.read(cuadro_path)) if cuadro_path else None
        sheet_data = root.find(NS + "sheetData")
        rows = {int(r.attrib["r"]): r for r in sheet_data.findall(NS + "row")}

        for pivot_table_def in root.findall(NS + "pivotTableDefinition"):
            root.remove(pivot_table_def)
        if resumen_rels is not None:
            for rel in list(resumen_rels):
                if rel.attrib.get("Type", "").endswith("/pivotTable"):
                    target = rel.attrib.get("Target", "")
                    removed_part = "xl/" + target.replace("../", "").lstrip("/")
                    removed_resumen_relationship_parts.add(removed_part)
                    removed_resumen_relationship_parts.add(removed_part.replace("pivotTables/", "pivotTables/_rels/") + ".rels")
                    resumen_rels.remove(rel)

        year, mes, etiqueta_mes = parse_etiqueta_mes(etiqueta_mes)
        total_col = None
        total_general_col = None
        month_total_cols = []

        for c in rows.get(4, []).findall(NS + "c"):
            valor = str(cell_value(c, shared)).strip()
            col_idx = split_ref(c.attrib["r"])[0]
            if valor == f"Total {etiqueta_mes}":
                total_col = col_idx
            elif valor == "Total general":
                total_general_col = col_idx
            elif valor.startswith("Total ") and valor not in {"Total general", "Total Var."}:
                month_total_cols.append(col_idx)

        if not total_general_col:
            SHARED_STRINGS_CONTEXT = None
            raise ValueError("No se encontro la columna 'Total general' en la fila 4 del comparativo.")

        if total_col and accion_mes == "agregar":
            total_col = None

        if total_col:
            start_col = total_col - 6
            previous_start_col = start_col - 7 if start_col > 8 else start_col
        else:
            if not month_total_cols:
                SHARED_STRINGS_CONTEXT = None
                raise ValueError("No se encontraron bloques de meses en la fila 4 del comparativo.")
            previous_total_col = max(col for col in month_total_cols if col < total_general_col)
            previous_start_col = previous_total_col - 6
            start_col = total_general_col
            total_col = start_col + 6
            shift_sheet_columns(root, total_general_col, 7)
            total_general_col += 7
            month_total_cols.append(total_col)

            rows = {int(r.attrib["r"]): r for r in sheet_data.findall(NS + "row")}
            row4 = rows[4]
            row5 = rows[5]
            for i in range(7):
                copy_style(find_cell(row4, previous_start_col + i, 4), get_or_create_cell(row4, start_col + i, 4))
                copy_style(find_cell(row5, previous_start_col + i, 5), get_or_create_cell(row5, start_col + i, 5))

            set_number(get_or_create_cell(row4, start_col, 4), excel_serial(date(year, mes, 1)))
            for i in range(1, 6):
                set_number(get_or_create_cell(row4, start_col + i, 4), None)
            set_text(get_or_create_cell(row4, total_col, 4), f"Total {etiqueta_mes}")

            for i, label in enumerate(["dev", "devqa", "qa", "", "prod", "all"]):
                set_text(get_or_create_cell(row5, start_col + i, 5), label)
            set_number(get_or_create_cell(row5, total_col, 5), None)

        month_total_cols = sorted(set(month_total_cols))
        entornos = ["dev", "devqa", "qa", "", "prod", "all"]
        variation_start_col = total_general_col + 3
        previous_label = str(cell_value(find_cell(rows[4], start_col - 1, 4), shared) or "").replace("Total ", "")
        variation_label = f"Variacion {previous_label}-{etiqueta_mes}"
        servicio_actual = None

        for r_idx in range(6, max(rows) + 1):
            row = rows.get(r_idx)
            if row is None:
                continue
            for i in range(7):
                copy_style(find_cell(row, previous_start_col + i, r_idx), get_or_create_cell(row, start_col + i, r_idx))
            a_cell = find_cell(row, 1, r_idx)
            etiqueta = cell_value(a_cell, shared)
            if etiqueta is None:
                continue
            etiqueta = str(etiqueta).strip()
            datos = None
            if etiqueta == "Total general":
                datos = {e: sum(por_servicio[s].get(e, 0.0) for s in por_servicio) for e in entornos}
            elif etiqueta in por_servicio:
                servicio_actual = etiqueta
                datos = por_servicio[etiqueta]
            elif servicio_actual and (servicio_actual, etiqueta) in por_plan:
                datos = por_plan[(servicio_actual, etiqueta)]

            if datos is None:
                continue

            total = 0.0
            for i, entorno in enumerate(entornos):
                valor = float(datos.get(entorno, 0.0))
                total += valor
                set_number(get_or_create_cell(row, start_col + i, r_idx), valor if (valor != 0 or etiqueta == "Total general") else None)
            set_number(get_or_create_cell(row, total_col, r_idx), total)
            total_general = sum(a_numero(cell_value(find_cell(row, col, r_idx), shared)) for col in month_total_cols)
            set_number(get_or_create_cell(row, total_general_col, r_idx), total_general)

        row4 = rows[4]
        row5 = rows[5]
        set_text(get_or_create_cell(row4, variation_start_col, 4), variation_label)
        set_text(get_or_create_cell(row4, variation_start_col + 6, 4), "Total Var.")
        for i, label in enumerate(["dev", "devqa", "qa", "", "prod", "all"]):
            set_text(get_or_create_cell(row5, variation_start_col + i, 5), label)
        set_number(get_or_create_cell(row5, variation_start_col + 6, 5), None)

        for r_idx in range(6, max(rows) + 1):
            row = rows.get(r_idx)
            if row is None or find_cell(row, 1, r_idx) is None:
                continue
            for i in range(7):
                new_ref = f"{index_to_col(start_col + i)}{r_idx}"
                old_ref = f"{index_to_col(previous_start_col + i)}{r_idx}"
                set_formula(
                    get_or_create_cell(row, variation_start_col + i, r_idx),
                    f'IF(ABS(N({new_ref})-N({old_ref}))<0.005,"-",N({new_ref})-N({old_ref}))',
                )

        normalize_ignorable_prefixes(root)
        if cuadro_root is not None:
            normalize_ignorable_prefixes(cuadro_root)

        with zipfile.ZipFile(ruta_salida, "w", zipfile.ZIP_DEFLATED) as zout:
            for rel in [
                rel
                for rel in list(rels)
                if rel.attrib.get("Type", "").endswith("/calcChain")
            ]:
                rels.remove(rel)
            content_types = ET.fromstring(zin.read("[Content_Types].xml"))
            for override in list(content_types):
                part_name = override.attrib.get("PartName")
                if part_name == "/xl/calcChain.xml" or part_name in {f"/{p}" for p in removed_resumen_relationship_parts}:
                    content_types.remove(override)
            for item in zin.infolist():
                if item.filename == "xl/calcChain.xml" or item.filename in removed_resumen_relationship_parts:
                    continue
                data = zin.read(item.filename)
                if item.filename == resumen_path:
                    data = ET.tostring(root, encoding="utf-8", xml_declaration=True)
                elif item.filename == cuadro_path and cuadro_root is not None:
                    data = ET.tostring(cuadro_root, encoding="utf-8", xml_declaration=True)
                elif item.filename == "xl/sharedStrings.xml":
                    data = ET.tostring(shared_root, encoding="utf-8", xml_declaration=True)
                elif item.filename == resumen_rels_path and resumen_rels is not None:
                    data = ET.tostring(resumen_rels, encoding="utf-8", xml_declaration=True)
                elif item.filename == "xl/_rels/workbook.xml.rels":
                    data = ET.tostring(rels, encoding="utf-8", xml_declaration=True)
                elif item.filename == "[Content_Types].xml":
                    data = ET.tostring(content_types, encoding="utf-8", xml_declaration=True)
                zout.writestr(item, data)

        SHARED_STRINGS_CONTEXT = None
        if postproceso_excel and not reparar_con_excel_si_disponible(ruta_salida, ruta_instances, etiqueta_mes, accion_mes):
            raise RuntimeError("Excel no pudo actualizar el comparativo. Cierra cualquier ventana de Excel abierta y vuelve a intentar.")


def main():
    parser = argparse.ArgumentParser(description="Genera archivos separados XM_Summary/XM_Instances y agrega el mes al comparativo.")
    parser.add_argument("--summary", required=True, help="XM-summary exportado sin tabla dinámica.")
    parser.add_argument("--instances", required=True, help="XM-Instances exportado sin tabla dinámica.")
    parser.add_argument("--comparativo", required=True, help="Comparativo base.")
    parser.add_argument("--mes", default="Jun-26", help="Mes a agregar. Ejemplo: Jul-26")
    parser.add_argument("--salida-summary", default="XM_Summary_tabla_auto.xlsx")
    parser.add_argument("--salida-instances", default="XM_Instances_tabla_auto.xlsx")
    parser.add_argument("--salida-comparativo", default="Comparativo_Instance_actualizado.xlsx")
    args = parser.parse_args()

    summary = consolidar_summary(args.summary)
    por_plan, por_servicio, servicios_planes = consolidar_instances(args.instances)
    generar_excels_tablas(args.salida_summary, args.salida_instances, summary, por_plan, por_servicio, servicios_planes)
    actualizar_comparativo_xml(
        args.comparativo,
        args.salida_comparativo,
        args.mes,
        por_plan,
        por_servicio,
        "auto",
        args.instances,
    )

    print("Archivos generados:")
    print(f"- {args.salida_summary}")
    print(f"- {args.salida_instances}")
    print(f"- {args.salida_comparativo}")


if __name__ == "__main__":
    main()
